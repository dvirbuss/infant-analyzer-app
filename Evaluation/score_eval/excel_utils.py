import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from sklearn.metrics import confusion_matrix
import excel_config as conf

def apply_borders(worksheet, cell_range):
    """Applies a thin border to all cells in the specified range."""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet[cell_range]:
        for cell in row:
            cell.border = thin_border

def apply_centering(worksheet, cell_range):
    """Centers text horizontally and vertically in the specified range."""
    center_align = Alignment(horizontal='center', vertical='center')
    for row in worksheet[cell_range]:
        for cell in row:
            cell.alignment = center_align

def apply_headline_format(worksheet, cell_range):
    """Bolds text and applies the headline color to the specified range."""
    headline_fill = PatternFill(start_color=conf.COLOR_HEADLINE, end_color=conf.COLOR_HEADLINE, fill_type="solid")
    bold_font = Font(bold=True)
    for row in worksheet[cell_range]:
        for cell in row:
            cell.fill = headline_fill
            cell.font = bold_font

def format_data_sheets(writer, sheet_names):
    """
    Sets the column width to 17 for the first 2 sheets (GT and model_pred).
    Colors the headline cells in light-mid grey.
    """
    grey_fill = PatternFill(start_color=conf.COLOR_HEADLINE, end_color=conf.COLOR_HEADLINE, fill_type="solid")
    
    for sheet_name in sheet_names:
        if sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for col in worksheet.columns:
                column_letter = col[0].column_letter
                worksheet.column_dimensions[column_letter].width = 17
            
            # Color the first row (headlines) grey
            for cell in worksheet[1]:
                cell.fill = grey_fill

from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

def format_analytics_sheet(writer, sheet_name, gt_df, pred_df):
    """
    Dynamically adjusts column widths, adds a confusion matrix, colors it,
    uses formulas to calculate metrics and confusion matrix from other sheets,
    and adds a pie chart seeded with actual calculated values.
    """
    if sheet_name not in writer.sheets:
        return
        
    worksheet = writer.sheets[sheet_name]
    
    # Calculate the dynamic range from the GT sheet
    gt_ws = writer.sheets[conf.SHEET_GT]
    max_row = gt_ws.max_row
    max_col_letter = get_column_letter(gt_ws.max_column)
    range_str = f"B2:{max_col_letter}{max_row}"
    
    gt_sheet = conf.SHEET_GT
    pred_sheet = conf.SHEET_PRED

    # Add formulas for accuracy, precision, recall, f1
    worksheet['A2'] = '=IFERROR(TRUNC((B5+C6)/(B5+C5+B6+C6), 2), 0)'
    worksheet['B2'] = '=IFERROR(TRUNC(B5/(B5+C5), 2), 0)'
    worksheet['C2'] = '=IFERROR(TRUNC(B5/(B5+B6), 2), 0)'
    worksheet['D2'] = '=IFERROR(TRUNC((2*B5)/(2*B5+C5+B6), 2), 0)'

    # Add confusion matrix layout
    worksheet['A4'].fill = PatternFill(start_color=conf.COLOR_BLACK, end_color=conf.COLOR_BLACK, fill_type="solid")
    worksheet['A5'] = "predicted-T"
    worksheet['A6'] = "predicted-F"
    worksheet['B4'] = "Real-T"
    worksheet['C4'] = "Real-F"
    
    # Dynamic formulas for confusion matrix
    worksheet['B5'] = f'=SUMPRODUCT(({gt_sheet}!{range_str}=1)*({pred_sheet}!{range_str}=1))' # TP
    worksheet['C5'] = f'=SUMPRODUCT(({gt_sheet}!{range_str}=0)*({gt_sheet}!{range_str}<>"")*({pred_sheet}!{range_str}=1))' # FP
    worksheet['B6'] = f'=SUMPRODUCT(({gt_sheet}!{range_str}=1)*({pred_sheet}!{range_str}=0)*({pred_sheet}!{range_str}<>""))' # FN
    worksheet['C6'] = f'=SUMPRODUCT(({gt_sheet}!{range_str}=0)*({gt_sheet}!{range_str}<>"")*({pred_sheet}!{range_str}=0)*({pred_sheet}!{range_str}<>""))' # TN

    # Color matrix cells
    light_green = PatternFill(start_color=conf.COLOR_LIGHT_GREEN, end_color=conf.COLOR_LIGHT_GREEN, fill_type="solid")
    light_red = PatternFill(start_color=conf.COLOR_LIGHT_RED, end_color=conf.COLOR_LIGHT_RED, fill_type="solid")
    
    worksheet['B5'].fill = light_green  # TP
    worksheet['C6'].fill = light_green  # TN
    worksheet['C5'].fill = light_red    # FP
    worksheet['B6'].fill = light_red    # FN

    # Apply Borders and Centering
    apply_borders(worksheet, conf.METRICS_TABLE)
    apply_borders(worksheet, conf.CONFUSION_MATRIX)
    
    apply_centering(worksheet, conf.METRICS_TABLE)
    apply_centering(worksheet, conf.CONFUSION_MATRIX)

    # Apply Headline formatting
    apply_headline_format(worksheet, conf.METRICS_HEADLINES)
    apply_headline_format(worksheet, conf.REAL_CONF_MAT)
    apply_headline_format(worksheet, conf.PRED_CONF_MAT)

    # Calculate actual values in Python for the chart cache
    import pandas as pd
    task_columns = [c for c in gt_df.columns if c != 'video_name']
    gt_values = gt_df[task_columns].values.flatten()
    pred_values = pred_df[task_columns].values.flatten()
    valid_indices = ~pd.isna(gt_values) & ~pd.isna(pred_values)
    
    if valid_indices.sum() > 0:
        y_true = gt_values[valid_indices].astype(int)
        y_pred = pred_values[valid_indices].astype(int)
        tn_val, fp_val, fn_val, tp_val = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    else:
        tn_val, fp_val, fn_val, tp_val = 0, 0, 0, 0

    # Add Pie Chart using explicit UNION references per the user's working formula
    from openpyxl.chart.series import Series, DataPoint
    from openpyxl.chart.data_source import NumRef, NumDataSource, AxDataSource, StrData, StrVal
    from openpyxl.chart.label import DataLabelList

    pie = PieChart()
    pie.title = 'Confusion Matrix Breakdown'

    ser = Series()
    sheet_esc = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
    
    # Use the exact formula structure requested: (B5, B6, C5, C6)
    union_formula = f"({sheet_esc}!$B$5,{sheet_esc}!$B$6,{sheet_esc}!$C$5,{sheet_esc}!$C$6)"
    ser.val = NumDataSource(numRef=NumRef(f=union_formula))

    # Categories matching the order: TP, FN, FP, TN
    cat_data = StrData()
    cat_data.ptCount = 4
    for i, v in enumerate(["TP", "FN", "FP", "TN"]):
        cat_data.pt.append(StrVal(idx=i, v=v))
    ser.cat = AxDataSource(strLit=cat_data)
    
    # Color chart slices (matching TP, FN, FP, TN order)
    # TP: Dark Green, FN: Dark Red (C00000), FP: Light Red (FF6666), TN: Light Green
    colors = ['00B050', 'C00000', 'FF6666', '92D050'] 
    for idx, color in enumerate(colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        ser.dPt.append(pt)

    pie.series.append(ser)
    
    # Show only Percentages and Category Names
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showSerName = False

    worksheet.add_chart(pie, conf.CHART_ANCHOR)

    # Dynamic column widths
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        
        # Skip hidden columns
        if column_letter in ['Z', 'AA']:
            continue
            
        for cell in col:
            try:
                # Handle formula string lengths safely
                val = cell.value
                if val is not None:
                    if str(val).startswith('='):
                        cell_length = 5 
                    else:
                        cell_length = len(str(val))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        
        adjusted_width = max_length + 3
        worksheet.column_dimensions[column_letter].width = adjusted_width
